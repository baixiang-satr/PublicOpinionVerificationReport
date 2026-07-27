"""Read URL candidates from TXT, CSV and standard OpenXML XLSX input files."""

from __future__ import annotations

import codecs
import csv
import io
import zipfile
from pathlib import Path

from src.domain.models import InputReadResult
from src.input.url_parser import build_url_tasks


SUPPORTED_INPUT_SUFFIXES = {".txt", ".csv", ".xlsx"}


class InputReadError(ValueError):
    """Raised when a user input file cannot safely be read as a URL source."""


def read_url_input(path: Path, sheet_name: str | None = None) -> InputReadResult:
    path = Path(path)
    if not path.is_file():
        raise InputReadError(f"Input file does not exist: {path}")
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_INPUT_SUFFIXES:
        raise InputReadError(f"Unsupported input file type: {suffix}")
    if suffix == ".txt":
        values = _read_text_values(path)
    elif suffix == ".csv":
        values = _read_csv_values(path)
    else:
        values = _read_xlsx_values(path, sheet_name)
    tasks, rejected = build_url_tasks(values)
    return InputReadResult(tuple(tasks), tuple(rejected), path)


def list_xlsx_sheets(path: Path) -> list[str]:
    workbook = _load_input_workbook(path)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _read_text_values(path: Path) -> list[str]:
    return _decode_bytes(path.read_bytes()).splitlines()


def _read_csv_values(path: Path) -> list[str]:
    text = _decode_bytes(path.read_bytes())
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel
    return [cell for row in csv.reader(io.StringIO(text), dialect) for cell in row]


def _read_xlsx_values(path: Path, sheet_name: str | None) -> list[str]:
    workbook = _load_input_workbook(path)
    try:
        selected_sheet = sheet_name or workbook.active.title
        if selected_sheet not in workbook.sheetnames:
            raise InputReadError(f"Worksheet does not exist: {selected_sheet}")
        sheet = workbook[selected_sheet]
        return [str(cell.value) for row in sheet.iter_rows(values_only=False) for cell in row if cell.value is not None]
    finally:
        workbook.close()


def _load_input_workbook(path: Path):
    if not zipfile.is_zipfile(path):
        raise InputReadError("Input XLSX must be a standard OpenXML workbook, not the fixed template.")
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise InputReadError("openpyxl is required to read XLSX input files.") from error
    try:
        return load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        raise InputReadError(f"Unable to read XLSX input: {path.name}") from error


def _decode_bytes(data: bytes) -> str:
    for encoding, marker in (("utf-8-sig", codecs.BOM_UTF8), ("utf-16", codecs.BOM_UTF16_LE), ("utf-16", codecs.BOM_UTF16_BE)):
        if data.startswith(marker):
            return data.decode(encoding)
    try:
        from chardet import detect

        detected = detect(data).get("encoding")
        if detected:
            return data.decode(detected)
    except ImportError:
        pass
    for encoding in ("utf-8", "gb18030", "utf-16"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise InputReadError("Unable to detect input text encoding.")
