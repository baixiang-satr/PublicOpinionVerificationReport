"""Import a previously delivered ``template.zip`` as a review session.

The user may upload an earlier deliverable and continue manual entry on top
of it: every non-empty business row of the eight fixed worksheets becomes a
:class:`RecordResult`, existing screenshot/attachment files referenced by the
workbook are re-linked, and the reconstructed records are persisted as a
regular ``job_checkpoint.json`` so the standard re-export pipeline applies
unchanged.

Reading the workbook uses ``openpyxl`` in read-only mode — the same library
already used for user-provided XLSX inputs.  Writing the fixed template is
never done here; export still goes through staging + ``OoxmlTemplateWriter``.
"""
from __future__ import annotations

from datetime import datetime
import hashlib
import json
from pathlib import Path
import zipfile

from src.domain.models import (
    AssetSet,
    PageData,
    RecordResult,
    RecordStatus,
    RouteDecision,
    UrlTask,
)
from src.domain.template_schema import SHEET_LAYOUTS, SHEET_ORDER, SheetLayout
from src.services.job_records import write_checkpoint
from src.utils.file_utils import atomic_replace
from src.utils.time_utils import DEFAULT_TIMEZONE


class TemplateZipImportError(ValueError):
    """Raised when the uploaded file is not a usable template.zip."""


class TemplateZipImporter:
    def __init__(self, output_root: Path) -> None:
        self._output_root = Path(output_root)

    def import_zip(self, zip_path: Path) -> Path:
        """Extract ``zip_path`` into a new job directory and return it."""

        zip_path = Path(zip_path)
        if not zip_path.is_file():
            raise TemplateZipImportError(f"文件不存在：{zip_path}")
        try:
            archive = zipfile.ZipFile(zip_path)
        except zipfile.BadZipFile as error:
            raise TemplateZipImportError("这不是一个有效的 ZIP 压缩包。") from error
        with archive:
            names = set(archive.namelist())
            if "template/template.xlsx" not in names:
                raise TemplateZipImportError(
                    "压缩包中找不到 template/template.xlsx，"
                    "请上传本工具生成的 template.zip。"
                )
            job_id = self._new_job_id(zip_path)
            job_dir = self._output_root / job_id
            staging_dir = job_dir / "staging"
            staging_dir.mkdir(parents=True, exist_ok=True)
            archive.extractall(staging_dir)

        template_dir = staging_dir / "template"
        records = self._read_records(template_dir / "template.xlsx", template_dir)
        if not records:
            raise TemplateZipImportError(
                "工作簿 8 张表的第 3 行起没有任何已填写内容，无法补录。"
            )
        write_checkpoint(job_dir, job_id, records)
        self._write_manifest(job_dir, zip_path, len(records))
        return job_dir

    # ── workbook reading ──
    def _read_records(self, workbook_path: Path, template_dir: Path) -> list[RecordResult]:
        try:
            from openpyxl import load_workbook
        except ImportError as error:  # pragma: no cover - dependency present
            raise TemplateZipImportError("缺少 openpyxl，无法读取工作簿。") from error
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        records: list[RecordResult] = []
        evidence_id = 0
        try:
            for sheet_name in SHEET_ORDER:
                if sheet_name not in workbook.sheetnames:
                    continue
                layout = SHEET_LAYOUTS[sheet_name]
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows(min_row=layout.data_start_row):
                    values = _row_business_values(row, layout)
                    if not any(values.values()):
                        continue
                    evidence_id += 1
                    records.append(
                        self._build_record(evidence_id, layout, values, template_dir)
                    )
        finally:
            workbook.close()
        return records

    def _build_record(
        self,
        evidence_id: int,
        layout: SheetLayout,
        values: dict[str, object],
        template_dir: Path,
    ) -> RecordResult:
        def cell(field: str) -> str:
            column = layout.field_columns.get(field)
            return _text(values.get(column)) if column else ""

        url = cell("url")
        platform = cell("platform")
        text_type = cell("text_type")
        published_raw = _text(values.get(layout.field_columns.get("published_at", "")))
        primary_name = _text(values.get(layout.primary_screenshot_column or ""))
        attachment_text = _text(values.get(layout.attachment_column or ""))

        assets = AssetSet()
        if primary_name:
            candidate = template_dir / primary_name
            if candidate.is_file():
                assets.page_screenshot = candidate
        for name in (part.strip() for part in attachment_text.split(",") if part.strip()):
            candidate = template_dir / name
            if candidate.is_file() and name != primary_name:
                assets.extra_attachments.append(candidate)

        status = (
            RecordStatus.NEEDS_REVIEW
            if self._missing_required(layout, values, assets)
            else RecordStatus.EXPORTED
        )
        return RecordResult(
            task=UrlTask(evidence_id, url, url),
            status=status,
            page=PageData(
                final_url=url or None,
                title=cell("title") or None,
                content_text=cell("content") or None,
                author_name=cell("author_name") or None,
                author_id=cell("author_id") or None,
                account_uin=cell("account_uin") or None,
                store_name=cell("store_name") or None,
                published_at=_parse_datetime(values.get(layout.field_columns.get("published_at", ""))),
                published_at_raw=published_raw or None,
            ),
            route=RouteDecision(
                sheet_name=layout.name,
                platform_value=platform,
                text_type=text_type or "正文",
            ),
            assets=assets,
        )

    @staticmethod
    def _missing_required(
        layout: SheetLayout,
        values: dict[str, object],
        assets: AssetSet,
    ) -> bool:
        for column in layout.required_columns:
            if column == layout.primary_screenshot_column:
                if assets.page_screenshot is None:
                    return True
            elif not _text(values.get(column)):
                return True
        return False

    # ── job dir ──
    def _new_job_id(self, zip_path: Path) -> str:
        stamp = datetime.now(DEFAULT_TIMEZONE).strftime("%Y%m%d-%H%M%S")
        digest = hashlib.sha256(
            f"{zip_path.resolve()}|{zip_path.stat().st_mtime_ns}".encode("utf-8")
        ).hexdigest()[:6]
        job_id = f"{stamp}-imported-{digest}"
        while (self._output_root / job_id).exists():
            digest = hashlib.sha256((digest + "x").encode("utf-8")).hexdigest()[:6]
            job_id = f"{stamp}-imported-{digest}"
        return job_id

    @staticmethod
    def _write_manifest(job_dir: Path, zip_path: Path, record_count: int) -> None:
        manifest = {
            "source_zip": str(zip_path),
            "imported_at": datetime.now(DEFAULT_TIMEZONE).isoformat(),
            "record_count": record_count,
        }
        target = job_dir / "import_manifest.json"
        temporary = target.with_suffix(".tmp")
        temporary.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        atomic_replace(temporary, target)


def _row_business_values(row, layout: SheetLayout) -> dict[str, object]:
    values: dict[str, object] = {}
    for cell in row:
        if cell.value is None:
            continue
        column = cell.column_letter
        if column in _layout_columns(layout):
            values[column] = cell.value
    return values


def _layout_columns(layout: SheetLayout) -> set[str]:
    columns = set(layout.field_columns.values())
    if layout.primary_screenshot_column:
        columns.add(layout.primary_screenshot_column)
    if layout.attachment_column:
        columns.add(layout.attachment_column)
    return columns


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if not isinstance(value, str) or not value.strip():
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(value.strip(), fmt)
        except ValueError:
            continue
    return None
