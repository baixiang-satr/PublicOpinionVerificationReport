"""Round-trip tests for importing a delivered template.zip into a review job."""
from __future__ import annotations

from pathlib import Path
import zipfile

import pytest

from src.domain.template_schema import SHEET_LAYOUTS, SHEET_ORDER
from src.services.checkpoint_store import CheckpointStore
from src.services.review_session import ReviewSession
from src.services.zip_import import TemplateZipImporter, TemplateZipImportError


def _build_template_zip(target: Path) -> Path:
    """Create a minimal but contract-shaped template.zip fixture."""

    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEET_ORDER:
        layout = SHEET_LAYOUTS[name]
        sheet = workbook.create_sheet(name)
        sheet.append(list(layout.headers))
        sheet.append(["示例"] * layout.column_count)
    weibo = workbook["微博博客"]
    weibo.append([
        "https://example.test/post/1",
        "昵称甲",
        "新浪_新浪微博_博客贴吧",
        "正文",
        "2026-07-01 12:00:00",
        "完整内容",
        "shot_001.png",
        "author_001.png,extra_001.png",
    ])
    weibo.append([
        "https://example.test/post/2",
        "",
        "新浪_新浪微博_博客贴吧",
        "正文",
        "",
        "",
        "",
        "",
    ])
    qun = workbook["群聊"]
    qun.append(["", "", "微信-群聊", "用户A", "", "群聊内容", "", "chat.png", ""])

    staging = target / "build"
    (staging / "template").mkdir(parents=True)
    workbook.save(staging / "template" / "template.xlsx")
    (staging / "template" / "shot_001.png").write_bytes(b"\x89PNG\r\n\x1a\nfake")
    (staging / "template" / "author_001.png").write_bytes(b"\x89PNG\r\n\x1a\nauthor")
    (staging / "template" / "extra_001.png").write_bytes(b"\x89PNG\r\n\x1a\nextra")
    (staging / "template" / "chat.png").write_bytes(b"\x89PNG\r\n\x1a\nfake")

    zip_path = target / "template.zip"
    with zipfile.ZipFile(zip_path, "w") as archive:
        for path in (staging / "template").iterdir():
            archive.write(path, f"template/{path.name}")
    return zip_path


def test_import_reconstructs_records_checkpoint_and_assets(tmp_path: Path) -> None:
    zip_path = _build_template_zip(tmp_path)
    importer = TemplateZipImporter(tmp_path / "output")
    job_dir = importer.import_zip(zip_path)

    assert job_dir.is_dir()
    assert (job_dir / "staging" / "template" / "template.xlsx").is_file()
    snapshot = CheckpointStore.load(job_dir / "job_checkpoint.json")
    assert len(snapshot.records) == 3

    first, second, third = snapshot.records
    # 证据编号按模板表序全局编号：群聊在微博博客之前
    assert first.task.evidence_id == 1
    assert first.route.sheet_name == "群聊"
    assert first.task.original_url == ""
    assert first.page.author_name == "用户A"
    assert first.assets.page_screenshot is not None

    assert second.task.original_url == "https://example.test/post/1"
    assert second.route.sheet_name == "微博博客"
    assert second.page.author_name == "昵称甲"
    assert second.page.published_at is not None
    assert second.assets.page_screenshot is not None
    assert second.assets.page_screenshot.name == "shot_001.png"
    assert second.assets.author_screenshot is not None
    assert second.assets.author_screenshot.name == "author_001.png"
    assert [path.name for path in second.assets.extra_attachments] == ["extra_001.png"]
    # 完整行 → 已导出；缺失行 → 待补录
    assert second.status.value == "exported"
    assert third.status.value == "needs_review"

    # 导入后可直接进入补录会话并保存人工修改
    session = ReviewSession.from_job_dir(job_dir)
    session.set_field(3, "author_name", "补录昵称")
    session.set_field(3, "platform", "新浪_新浪微博_博客贴吧")
    reloaded = ReviewSession.from_job_dir(job_dir)
    views = {v.field: v for v in reloaded.field_views(3)}
    assert views["author_name"].value == "补录昵称"
    assert views["platform"].value == "新浪_新浪微博_博客贴吧"


def test_import_rejects_non_zip_and_missing_workbook(tmp_path: Path) -> None:
    importer = TemplateZipImporter(tmp_path / "output")
    bad = tmp_path / "bad.zip"
    bad.write_text("not a zip", encoding="utf-8")
    with pytest.raises(TemplateZipImportError):
        importer.import_zip(bad)

    empty = tmp_path / "empty.zip"
    with zipfile.ZipFile(empty, "w") as archive:
        archive.writestr("readme.txt", "hello")
    with pytest.raises(TemplateZipImportError):
        importer.import_zip(empty)


def test_import_swapped_sheet_restores_homepage_to_author_slot(tmp_path: Path) -> None:
    """图文视频对调表：H 列还原为个人主页截图资产，I 列首项还原为内容页截图。"""

    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEET_ORDER:
        layout = SHEET_LAYOUTS[name]
        sheet = workbook.create_sheet(name)
        sheet.append(list(layout.headers))
        sheet.append(["示例"] * layout.column_count)
    video = workbook["图文视频"]
    video.append([
        "https://v.douyin.com/abc/",
        "85741182891",
        "昵称甲",
        "字节跳动_抖音_图文视频",
        "正文",
        "2026-07-01 12:00:00",
        "完整内容",
        "001主页.jpg",
        "001.jpg,001_extra.png",
    ])

    staging = tmp_path / "build"
    (staging / "template").mkdir(parents=True)
    workbook.save(staging / "template" / "template.xlsx")
    for name in ("001.jpg", "001主页.jpg", "001_extra.png"):
        (staging / "template" / name).write_bytes(b"\x89PNG\r\n\x1a\nfake")
    zip_path = tmp_path / "template.zip"
    with zipfile.ZipFile(zip_path, "w") as archive:
        for path in (staging / "template").iterdir():
            archive.write(path, f"template/{path.name}")

    importer = TemplateZipImporter(tmp_path / "output")
    job_dir = importer.import_zip(zip_path)

    snapshot = CheckpointStore.load(job_dir / "job_checkpoint.json")
    assert len(snapshot.records) == 1
    record = snapshot.records[0]
    assert record.route is not None and record.route.sheet_name == "图文视频"
    assert record.assets.author_screenshot is not None
    assert record.assets.author_screenshot.name == "001主页.jpg"
    assert record.assets.page_screenshot is not None
    assert record.assets.page_screenshot.name == "001.jpg"
    assert [path.name for path in record.assets.extra_attachments] == ["001_extra.png"]
    # 主截图列（个人主页截图）已满足 → 不判缺失
    assert record.status.value == "exported"

    session = ReviewSession.from_job_dir(job_dir)
    assert session.primary_screenshot_name(record) == "001主页.jpg"
    assert session.content_screenshot_name(record) == "001.jpg"
